# import the required modules
import tabula
from openpyxl import load_workbook
import pandas as pd
from pandas.api.types import is_string_dtype
import sqlite3

# read a pdf file
value = tabula.read_pdf("Tabela-Materias-Unifesp.pdf", pages="all", encoding="utf-8")[0]

# convert PDF into CSV format
tabula.convert_into(
    "Tabela-Materias-Unifesp.pdf",
    "Materias-Unifesp-test.csv",
    output_format="csv",
    pages="all",
)


########### DEFINE TABLE AS DATAFRAME WITH PANDAS ################

# Input
data_file = "csvFiles/Materias-Unifesp-1.csv"

# Delimiter
data_file_delimiter = ","

# The max column count a line in the file could have
largest_column_count = 0

# Loop the data lines
with open(data_file, "r") as temp_f:
    # Read the lines
    lines = temp_f.readlines()

    for l in lines:
        # Count the column count for the current line
        column_count = len(l.split(data_file_delimiter)) + 1

        # Set the new most column count
        largest_column_count = (
            column_count
            if largest_column_count < column_count
            else largest_column_count
        )

# Generate column names (will be 0, 1, 2, ..., largest_column_count - 1)
column_names = [i for i in range(0, largest_column_count)]

# Read csv
df = pd.read_csv(
    data_file, header=None, delimiter=data_file_delimiter, names=column_names
)

# Define the header of the 4-th column
value = "Semestre"

# transform column of index 4 in string type
df[4] = df[4].astype(str)

# check if it's string type
print(is_string_dtype(df[4]))

# define the HEADER of the column of index 4
df.at[0, 4] = value

################ COMPARASION WITH THE GRADE HOUR FROM THE SEMESTER ################

# load the file to the
wb = load_workbook(filename="csvFiles/Matriz horária 1-2022.xlsx")

print("------ Loading sheet into wb ------")
# load the first sheet of the workbook
sheet = wb["Matriz Horária para 1º semestre"]

print("------ saving uc_names into list ------")
# save all the UCs names in a list
uc_names = df[0].tolist()[1:]

print("------ saving uc_names into dictionary ------")
# save all the UCs names into a dictionary and set them to false
uc_names_dict = dict.fromkeys(df[0][1:], False)

# Function to compare the values from the workbook file with the dataframe
def which_semester(uc_name):
    print("------ inside function which_semester ------")
    for i in range(1, sheet.max_row + 1):
        for j in range(sheet.max_column):
            if sheet[i][j].value == uc_name:
                for name in uc_names_dict:
                    if uc_name == name:
                        uc_names_dict.update({name: True})
                return


print("------ comparasion for loop ------")
# driver code to compare each UC name from our list of subjects
for i in range(len(uc_names)):
    which_semester(uc_names[i])


print("------ Inserting into df IMPAR/PAR ------")
# Insert into the Materias-Unifesp-1.csv file if the semester is IMPAR or PAR
# based on the uc_names_dict dictionary that has True if they found the uc in the Matriz table
for k in range(1, len(df[4])):
    for value in uc_names_dict:
        if uc_names_dict[value] == True and value == df[0][k]:
            df.at[k, 4] = "IMPAR"

# print the full dataframe
print(df)

# save the dataframe into an .csv file
# df.to_csv('Materias-Unifesp-Atualizada.csv')

